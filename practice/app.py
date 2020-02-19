import openpyxl as xl
from openpyxl.chart import BarChart,Reference
wb = xl.load_workbook('C:/Users/AFH轰炸机/Desktop/python/pytest/practice/1.xlsx')
sheet=wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)#也可以这样表示
print(sheet.max_row)
for row in range(2,sheet.max_row+1):
    cell = sheet.cell(row,3)
    print(cell.value)
    correct_cell_price = cell.value * 0.9
    correct_cell = sheet.cell(row, 4)
    correct_cell.value = correct_cell_price
wb.save('2.xlsx')
